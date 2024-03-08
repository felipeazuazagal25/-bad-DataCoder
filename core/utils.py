from fpdf import FPDF
from .models import Test, Question, QuestionRubric, Student, StudentTest, StudentTestQuestion
import random
import random, string
import openpyxl as opxl
import numpy as np


def generate_test(email,test):
    tests=Question.objects.filter(test__testName = test.testName)
    test_questionsNumber_dict=tests.values('questionNumber').distinct()
    test_questionsNumber_list=[question['questionNumber']for question in test_questionsNumber_dict]
    questions = []
    for i in range(len(test_questionsNumber_list)):
        test_questionVersion_dict=Question.objects.filter(test__testName = test, questionNumber = test_questionsNumber_list[i])
        test_questionVersion_dict=test_questionVersion_dict.values('questionVersion').distinct()
        test_questionVersion_list=[question['questionVersion'] for question in test_questionVersion_dict]
        version_selected = random.choice(test_questionVersion_list) # Random selection of questions
        question_selected = Question.objects.filter(test__testName = test, questionNumber = test_questionsNumber_list[i],questionVersion = version_selected).first()
        questions.append(question_selected)
    for question in questions:
        student_test = StudentTest.objects.filter(student__userModel__email = email).first()
        student_test_question = StudentTestQuestion(studentTest=student_test, questionTest=question)
        student_test_question.save()
    return 0




def generate_code(email,test):
    code = ''.join(random.choice(string.ascii_uppercase + string.ascii_lowercase + string.digits) for _ in range(16))
    student=Student.objects.filter(userModel__email = email).first()
    student_test=StudentTest(student=student,
                                 test=test,
                                 code=code)
    student_test.save()
    return code

def get_code_from_email(email, slug):
    code = StudentTest.objects.filter(student__userModel__email = email, test__testName = slug).first()
    return code.code



def get_version_from_email_question(email, test_name, question_number):
    temp = StudentTestQuestion.objects.filter(studentTest__student__userModel__email = email,
                                              questionTest__test__testName = test_name).all()
    print(temp)
    student_test = StudentTestQuestion.objects.filter(studentTest__student__userModel__email = email,
                                                      questionTest__test__testName = test_name,
                                                      questionTest__questionNumber = question_number).first()
    
    

    return student_test.questionTest.questionVersion


def grade_test(email,test_name,answers):
    grade = 1

    # # Open the answers file

    # wb = opxl.load_workbook(answers,data_only=False)
    # wb_data = opxl.load_workbook(answers,data_only=True)

    # # Iterate trough all the answers
    # questions_numbers = wb.sheetnames
    # for question_number in questions_numbers:
    #     fs = wb[question_number]
    #     fs_data = wb_data[question_number]
    #     fs_count_row = fs.max_row 
    #     fs_count_col = fs.max_column
        

    #     print(email)
    #     print(test_name)
    #     print(question_number)
        
    #     version = get_version_from_email_question(email,test_name,question_number)


    #     questions_rubric = QuestionRubric.objects.filter(question__test__testName = test_name,
    #                                                      question__questionNumber = question_number,
    #                                                      question__questionVersion = version).all()
    #     score_for = np.array([])
    #     for rubric in questions_rubric:
    #         cell= fs.cell(column=rubric.cellColumn,row=rubric.cellRow)
    #         cell_data= fs_data.cell(column=rubric.cellColumn,row=rubric.cellRow)
            
    #         # Check if it is formula
    #         if (cell.value[0]=='=') or (cell.value[0]=='+') or (cell.value[0]=='-'):
    #             score_for = np.append(score_for,[1])
    #         else:
    #             score_for = np.append(score_for,[0])

    #         # Check if the values coincide

    #     final_score =sum(score_for)/score_for.size

    # grade = 1 + final_score*6
    return 4.5



