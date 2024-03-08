from django.shortcuts import render
from django.shortcuts import redirect
from django.contrib.auth.decorators import login_required
from .forms import CourseForm, TestForm, TestQuestionsMassiveForm, StudentForm, ValidateCodeForm, TestDateForm, AnswersForm
from .models import Course, Test , Question , QuestionRubric, TestDate, Student, StudentTest, StudentTestQuestion
from authentication.models import User
import pandas as pd
from .utils import generate_test, generate_code, get_code_from_email, grade_test
from django.core.mail import send_mail
from .decorators import teacher_required, student_required
import json
from django.utils import timezone
import zipfile, os, io
import openpyxl as opxl




@login_required(login_url="auth/login/")
def course_creation_view(request):
    form = CourseForm(request.POST or None)
    if request.method == "POST":
        if form.is_valid():
            form.save()
            print('Form saved!')
            return redirect('core:courses_view')
        else:
            return render(request, "creations/course_creation.html", {"form": form})
    return render(request, "creations/course_creation.html", {"form": form})


@login_required(login_url="auth/login/")
def test_creation_view(request):
    form = TestForm(request.POST or None)
    if request.method == "POST":
        if form.is_valid():
            form.save()
            print('Form saved!')
            return redirect('core:courses_view')
        else:
            return render(request, "creations/question_test_creation.html", {"form": form})
    return render(request, "creations/question_test_creation.html", {"form": form})


@login_required(login_url="auth/login/")

def massive_test_question_creation_view(request):
    form = TestQuestionsMassiveForm(request.POST or None)
    if request.method == "POST":

        # Getting the files from the form
        questions = request.FILES['questions']
        rubric = request.FILES['rubric']
        rubric_paths = []

        # Reading the rubric files and getting their paths
        with zipfile.ZipFile(rubric, 'r') as zip_ref:
            temp_dir = 'rubrics/'
            zip_ref.extractall(temp_dir)
            for root, dirs, files in os.walk(temp_dir):
                for file in files:
                    if not file.startswith('._'):
                        rubric_paths.append(file)

        # Uploading the questions
        df_questions = pd.read_excel(questions)
        test = Test(testName = form['testName'].value())
        test.save()
        #print('Test created!')
        for i in range(len(df_questions.index)):
            question_test= Question(
                test=test,
                questionNumber=df_questions['Pregunta'].iloc[i],
                questionVersion=df_questions['Versión'].iloc[i],
                questionText=df_questions['Texto'].iloc[i],
                questionScore=int(df_questions['Puntaje'].iloc[i])
                )
            question_test.save()
            # print('Question ' + str(i)  + ' saved!')
            
            # Create the rubric for each question
            question_number = df_questions['Pregunta'].iloc[i]
            question_version = df_questions['Versión'].iloc[i]

            print('__________________')
            print('Question: ',question_number)
            print('Version: ',question_version)
            print('__________________')
            wb = opxl.load_workbook(temp_dir + question_version + '.xlsx',data_only=False)
            wb_data = opxl.load_workbook(temp_dir + question_version + '.xlsx',data_only=True)

            # Go to the respective sheet according to the question
            
            fs = wb[question_number]
            fs_data = wb_data[question_number]
            fs_count_row = fs.max_row 
            fs_count_col = fs.max_column

            for row in range(1,fs_count_row + 1):
                for column in range(1,fs_count_col+1):
                    cell = fs.cell(column=column,row=row)
                    cell_data = fs_data.cell(column=column,row=row)
                    bgColor = cell.fill.start_color.index
                    if(bgColor == 'FFFF0000'):
                        question_rubric = QuestionRubric(question = question_test,
                                                            cellRow = row,
                                                            cellColumn = column,
                                                            cellFormula = cell.value,
                                                            cellValue = cell_data.value)
                        question_rubric.save()
                        print(question_rubric)
        
            


        return redirect('core:tests_view')
    return render(request, "creations/question_test_massive_creation.html", {"form": form})


@login_required(login_url="auth/login/")
def delete_test(request,testName):
    test = Test.objects.filter(testName = testName).first()
    test.delete()
    return redirect('core:tests_view')


# Corregir student -> is_student flag
@login_required(login_url="auth/login/")
def student_creation_view(request):
    form = StudentForm(request.POST or None)
    courses = Course.objects.all()
    if request.method == "POST":
        if form.is_valid():
            cd = form.cleaned_data
            print(cd['email'].split('@')[0])
            user = User(username=cd['email'], email = cd['email'], is_student = True) #.split('@')[0] if I want to just have an email
            user.set_password(cd['password'])
            user.save()
            student = Student(userModel = user,course = cd['course'])
            student.save()
            print('Form saved!')
            return redirect('core:courses_view')
        else:
            return render(request, "creations/student_creation.html", {"form": form, 'courses' : courses})
    return render(request, "creations/student_creation.html", {"form": form, 'courses' : courses})


@login_required(login_url="auth/login/")
def questions_creation_form(request):
    return None

@login_required(login_url="auth/login/")
def student_creation_form(request):
    return None


@login_required(login_url="auth/login/")
def courses_view(request):
    load_template = request.path.split('/')[-1]
    courses = Course.objects.all()
    context = {'courses': courses, 'segment':load_template}
    return render(request,'views/courses.html',context)



@login_required(login_url="auth/login/")
def test_view(request):
    tests = Test.objects.all()
    tests = tests.values('testName')
    tests = [test['testName'] for test in tests]
    print(tests)
    questions = {}
    for test in tests:
        question=Question.objects.filter(test__testName = test)
        question=len(set([i['questionNumber'] for i in question.values('questionNumber')]))
        print(question)
        questions[test] = question
    #print(questions['Test1'])
    context = {'tests': tests, 'questions': questions}
    #print(context['questions'])
    return render(request,'views/tests.html',context)



@login_required(login_url="auth/login/")
def questions_view(request, slug):
    questions={}
    questionNumbers = []
    result = {}
    context={}
    if (Question.objects.filter(test__testName = slug)):
        context['slug']=slug
        questions = Question.objects.filter(test__testName = slug)
        for question in questions:
            questionNumbers.append(question.questionNumber)
        questionNumbers = list(set(questionNumbers))
        for num in questionNumbers:
            pregunta_dic = {}
            preguntas = Question.objects.filter(test__testName = slug, questionNumber = num)
            for version in preguntas:
                pregunta_dic[version.questionVersion] = version.questionText
            result[num]=pregunta_dic
    print(json.dumps(result,sort_keys=True, indent=4))
    result = dict(sorted(result.items()))
    context['questions'] = questions
    context['result'] = result
    return render(request,'views/questions.html',context)



def generate_test_view(request,slug):
    code_form=ValidateCodeForm(request.POST or None)
    answers_form = AnswersForm(request.POST or None)
    context={'code_form':code_form,
             'answers_form':answers_form,
             'slug':slug}

    # Getting information
    email = request.user.email
    test = Test.objects.filter(testName = slug).first()
    studentTest = StudentTest.objects.filter(student__userModel__email = email, test__testName = slug).first()
    if studentTest.status <= 2:
        if request.method == 'POST':

            # Code form

            if code_form.is_valid():
                # Clean data en get code from email
                cd = code_form.cleaned_data
                code=get_code_from_email(email,slug)
                if cd['code'] == code:

                    # Get the question from the database
                    questions_student = StudentTestQuestion.objects.filter(studentTest__student__userModel__email = email)
                    print('Cantidad de Preguntas:',questions_student.count())
                    if questions_student.count() == 0:
                        generate_test(email,test)

                    # Clean questions wihtout versions
                    question_student_clean = {}
                    for question in questions_student:
                        question_student_clean[question.questionTest.questionNumber] = question.questionTest.questionText
                    context['test'] = question_student_clean

                    # Save the datetime when initiated
                    if studentTest.takeDateStart is None:
                        studentTest.takeDateStart = timezone.now()
                        studentTest.status = 2
                        studentTest.save()
                    print(studentTest.takeDateStart)
                    context['remaining_time'] = 0
                else:
                    context['error_with_code'] = 'error_with_code'

            # Checking if the student surpass the limit of time

            if studentTest.takeDateStart is not None and studentTest.status <= 2:
                diff = timezone.now() - studentTest.takeDateStart
                context['remaining_time'] = int(diff.total_seconds())
                if diff.total_seconds() > 30*60:
                    # Remove the test from the html
                    print('Está dentro del if.')
                    context['test'] = None
                    context['remaining_time'] = None
                    context['end_test'] = 'endTest'

                    # Update the model with new info
                    studentTest.takeDateEnd = timezone.now()
                    studentTest.status = 3
                    studentTest.save()

                    return render(request,'views/generate_test.html',context)
            

            # If method == POST, check if there's any files.

            if len(request.FILES) != 0:
                answers = request.FILES['answers_excel']
                print('Got the answers')
                # Review test inmediately
                studentTest.grade = grade_test(email,slug,answers)
                studentTest.status = 4
                studentTest.save()
                
    elif studentTest.status == 4:
        context['test_taken'] = 'test_taken'

    return render(request,'views/generate_test.html',context)


def generate_practice(request):
    return None

@login_required(login_url="auth/login/")
def send_email(request,slug):
    from_email = 'felipeazuazagal25@gmail.com'
    students = Student.objects.all()
    emails = students.values('email').distinct() # Here the distinct shouldnt exist
    #print(emails)
    emails = [value['email'] for value in emails]    
    subject = 'Code for the test'
    base1 = 'Prepare for your test: '
    base2 = '\n Your code is: '
    for email in emails:
        code = generate_code(email,slug)
        message = base1 + slug + base2 + code
        send_mail(subject, message, from_email, [email])
    return redirect('core:tests_view')


@login_required(login_url="auth/login/")
def request_code(request,slug):
    context = {}
    from_email = 'felipeazuazagal25@gmail.com'
    #students = Student.objects.all()
    #emails = students.values('email').distinct() # Here the distinct shouldnt exist
    #print(emails)
    email = request.user.email
    subject = 'Código para rendir ' + slug
    base1 = 'Este es el código para rendir la prueba: ' + slug
    base2 = '\n Tu código es: '
    test = Test.objects.filter(testName=slug).first()
    code = StudentTest.objects.filter(student__userModel__email = email, test__testName = slug).first()
    code = code.code if code is not None else None
    if code is None:
        code = generate_code(email,test)
    message = base1+ base2 + code
    send_mail(subject, message, from_email, [email])
    context['msg'] = 'Tu código fue enviado con éxito. Revisa tu bandeja de entrada e ingresa el código para poder rendir la prueba.'
    url = '/core/generate_test/'+slug
    print(url)
    return redirect(url)




@login_required(login_url="auth/login/")
def test_date_creation_view(request):
    form = TestDateForm(request.POST or None)
    context = {}
    if request.method == "POST":
        if form.is_valid():
            testobject = Test.objects.filter(id=form['test'].value()).first()
            courseobject = Course.objects.filter(id=form['course'].value()).first()
            testdate = TestDate(test = testobject,course = courseobject,testdate = form['testdate'].value())
            try:
                testdate.save()
                print('Form saved!')
                students = Student.objects.filter(course = courseobject).all()
                for student in students:
                    code = generate_code(student.userModel.email,testobject)
                    print('Code Generated for one Student: ' + code)

                return redirect('core:tests_view')
            except:
                error = 'Prueba ya tiene fecha asignada. Intente con otra prueba/curso.'
                context['error'] = error
            
            
        else:
            context['form'] = form
            #return render(request, "creations/test_date_creation.html", context)
    context['form'] = form
    return render(request, "creations/test_date_creation.html", context)




@login_required(login_url="auth/login/")
@student_required
def tests_student_view(request):
    student = Student.objects.filter(userModel__username = request.user).first()
    tests = TestDate.objects.filter(course = student.course)
    context = {}
    result = {}
    for test in tests:
        student_test = StudentTest.objects.filter(student__userModel__email = request.user.email, test__testName = test.test.testName).first()
        info_test = {}
        info_test['num_questions'] = Question.objects.filter(test = test.test).values('questionNumber').distinct().count()
        info_test['lim_date'] = test.testdate.strftime('%m/%d/%y')
        info_test['lim_time'] = test.testdate.strftime('%H:%M')
        info_test['take_date'] = '-' if student_test.takeDateStart is None else student_test.takeDateStart.strftime('%m/%d/%y')
        info_test['take_time'] = '-' if student_test.takeDateStart is None else student_test.takeDateStart.strftime('%H:%M')
        info_test['status'] = student_test.get_status_display
        info_test['grade'] = '-' if student_test.grade is None else student_test.grade
        result[test.test.testName] = info_test
    print(result)
    context['result'] = result
    return render(request, 'views/tests_student.html',context)



@login_required(login_url="auth/login/")
@student_required
def demo_contenidos_view(request):
    context = {}
    return render(request, 'views/demo_contenidos.html',context)