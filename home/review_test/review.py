import openpyxl as opxl
import numpy as np

def review_file(file_dir,rubric):
    ans_ind,ans_val,ans_data = rubric
    test = opxl.load_workbook(file_dir,data_only=False)
    test_data = opxl.load_workbook(file_dir,data_only=True)
    fs = test.active # change it for specific name (try method, otherwise, go the active sheet)
    fs_data = test_data.active

    
    #Get the basic information
    student_email = [fs_data.cell(row = 7,column = 3).value]


    score_for = [] # create it with a fixed amount of questions
    score_dat = []

    for row,column in ans_ind:
        cell= fs.cell(column=column,row=row)
        cell_data= fs_data.cell(column=column,row=row)
        
        # Check if it is formula
        if (cell.value[0]=='=') or (cell.value[0]=='+') or (cell.value[0]=='-'):
            score_for.append(0.5)
        else:
            score_for.append(0)

        # Check if the values coincide
        current_cell = [i for i, j in enumerate(ans_ind) if j == [row,column]][0]

        if (cell_data.value == ans_data[current_cell]):
            score_dat.append(0.5)
        else:
            score_dat.append(0)

    score_final = score_for + score_dat

    final_data = student_email + score_final
    
    return final_data