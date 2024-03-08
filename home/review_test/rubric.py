import pandas as pd

import openpyxl as opxl

def create_rubric(file_dir):
    
    wb = opxl.load_workbook(file_dir,data_only=False)
    wb_data = opxl.load_workbook(file_dir,data_only=True)
    fs = wb.active # change it for specific name (try method, otherwise, go the active sheet)
    fs_data = wb_data.active
    fs_count_row = fs.max_row 
    fs_count_col = fs.max_column

    ans_ind = []
    ans_val = []
    ans_data = []

    for row in range(1,fs_count_row + 1):
        for column in range(1,fs_count_col+1):
            cell = fs.cell(column=column,row=row)
            cell_data = fs_data.cell(column=column,row=row)
            bgColor = cell.fill.start_color.index
            if(bgColor == 'FFFF0000'):
                ans_ind.append([row, column])
                ans_val.append(cell.value)
                ans_data.append(cell_data.value)

    # ans_group = split_answers(ans_ind)
    
    return[ans_ind,ans_val,ans_data]